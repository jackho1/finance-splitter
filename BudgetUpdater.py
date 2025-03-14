from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from datetime import datetime, timedelta
from copy import copy
import re
import os
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
import updateMyBuckets
from config import TRANSACTION_DIRECTORY

class BudgetUpdater:
    """ Updates the "Budget", "Total Balance", and "Jacks Buckets" sheets with recent transactions"""
    
    def __init__(self, file_path, workbook_name, one_drive_path=None, verbose=True):
        # Initialize basic parameters
        self.file_path = file_path
        self.workbook_name = workbook_name
        self.one_drive_path = one_drive_path
        self.verbose = verbose
        
        # Calculate previous and current month
        self.current_date = datetime.now()
        prev_month = self.current_date - timedelta(days=self.current_date.day)
        self.prev_month = prev_month.strftime("%B")
        self.prev_month_num = prev_month.month
        self.prev_year = prev_month.year
        self.current_month_num = self.current_date.month
        self.current_year = self.current_date.year
        
        # Load workbook
        self.wb = load_workbook(file_path, keep_vba=True, data_only=False)

    def _log(self, message, is_error=False):
        """Prints messages if verbose is True or if it's an error."""
        if is_error or self.verbose:
            print(message)

    def convert_previous_month_to_values(self):
        """Converts formulas to values and updates bold formatting for month rows in Total Balance sheet."""
        sheet_name = "Total Balance"
        if sheet_name not in self.wb.sheetnames:
            self._log(f"⚠ Sheet '{sheet_name}' not found.", is_error=True)
            return

        sheet = self.wb[sheet_name]
        self._log(f"\n🔹 Processing '{sheet_name}' for {self.prev_month}")

        try:
            wb_data = load_workbook(self.file_path, data_only=True, keep_vba=True)
            sheet_data = wb_data[sheet_name]
        except Exception as e:
            self._log(f"❌ Error loading workbook data: {str(e)}", is_error=True)
            return

        prev_month_row = None
        current_month_row = None

        # Scan for previous and current month rows
        for row in range(1, sheet.max_row + 1):
            cell = sheet[f'A{row}']
            if not isinstance(cell.value, datetime):
                continue

            cell_date = cell.value
            if cell_date.month == self.prev_month_num and cell_date.year == self.prev_year:
                prev_month_row = row
            elif cell_date.month == self.current_month_num and cell_date.year == self.current_year:
                current_month_row = row

            # Process previous month's row if found
            if prev_month_row and row == prev_month_row:
                self._log(f"✅ Found previous month row for {cell_date.strftime('%d/%m/%Y')}")
                
                # Convert formulas and unbold row
                for col in range(1, sheet.max_column + 1):  # Include column A
                    target = sheet.cell(row=row, column=col)
                    source = sheet_data.cell(row=row, column=col)
                    
                    # Convert formula to value
                    if target.data_type == 'f':
                        try:
                            if source.value is not None:
                                target.value = source.value
                                target.data_type = 'n' if isinstance(source.value, (int, float)) else 's'
                                self._log(f"🔄 Converted cell {get_column_letter(col)}{row}")
                            else:
                                self._log(f"⚠ No value for cell {get_column_letter(col)}{row}")
                        except Exception as e:
                            self._log(f"❌ Error in cell {get_column_letter(col)}{row}: {str(e)}", is_error=True)
                    
                    # Unbold cell
                    target.font = Font(bold=False)
                
                self._log(f"✅ Completed previous month row processing (unbolded)")

        # Bold current month's row if found
        if current_month_row:
            self._log(f"✅ Found current month row at A{current_month_row}")
            for col in range(1, sheet.max_column + 1):  # Include date column for current month
                cell = sheet.cell(row=current_month_row, column=col)
                cell.font = Font(bold=True)
            self._log(f"✅ Bolded current month row")
        else:
            self._log(f"⚠ No current month row found")

        if not prev_month_row:
            self._log(f"⚠ No previous month row found for {self.prev_month} {self.prev_year}")

    def update_formulas(self):
        """Updates monthly references in Budget sheet formulas."""
        if 'Budget' not in self.wb.sheetnames:
            self._log("⚠ Budget sheet not found.", is_error=True)
            return

        sheet = self.wb['Budget']
        self._log(f"\n🔹 Updating formulas in Budget sheet")
        current_month = datetime.now().strftime('%B')

        # Use a set for faster lookup of formula cells
        formula_cells = set()
        
        # First pass: identify cells with formulas
        for row in sheet.iter_rows():
            for cell in row:
                if self._has_formula(cell):
                    formula_cells.add(cell)

        # Second pass: update formulas
        for cell in formula_cells:
            try:
                original_formula = self._get_formula(cell)
                self._log(f"🔍 Checking if formula requires an update in {cell.coordinate}: {original_formula}")
                
                new_formula = self._update_formula(original_formula, current_month)
                
                if original_formula != new_formula:
                    self._log(f"📌 Updating {cell.coordinate}: {original_formula} → {new_formula}")
                    self._update_cell_value(cell, new_formula)
            except Exception as e:
                self._log(f"❌ Error updating {cell.coordinate}: {str(e)}", is_error=True)
                continue

    def _update_formula(self, formula, current_month):
        """Helper method to update formula references."""
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December']
        
        # Compile pattern once for better performance
        month_pattern = re.compile(r'\[([^\]]*?)\](' + '|'.join(months) + r')\b')
        
        def replace_match(match):
            workbook_ref = match.group(1)
            old_month = match.group(2)
            
            # Simplify workbook reference handling
            old_ref = f'[{workbook_ref}]{old_month}'
            new_ref = f"'[{self.workbook_name}]{current_month}'"
            return new_ref

        return month_pattern.sub(replace_match, formula)

    def _has_formula(self, cell):
        """Helper method to check if a cell contains a formula."""
        return (
            isinstance(cell.value, str) and cell.value.startswith("=") or
            isinstance(cell.value, ArrayFormula)
        )

    def _get_formula(self, cell):
        """Helper method to extract formula text from a cell."""
        if isinstance(cell.value, ArrayFormula):
            return cell.value.text if hasattr(cell.value, 'text') else cell.value
        return cell.value

    def _update_cell_value(self, cell, new_formula):
        """Helper method to update cell value while preserving formula type."""
        if isinstance(cell.value, ArrayFormula):
            cell.value = ArrayFormula(ref=cell.coordinate, text=new_formula)
        else:
            cell.value = new_formula

    # Add this at the beginning of your class or method to define and register custom styles
    def setup_styles(self, workbook):
        style_names = [style.name for style in workbook._named_styles]
        
        # Date style using Excel's built-in date format
        if "date_style" not in style_names:
            date_style = NamedStyle(name="date_style", number_format='D/MM/YYYY')
            date_style.alignment = Alignment(horizontal='center', vertical='center')
            workbook.add_named_style(date_style)
        self.date_style = "date_style"

        # Currency style using standard Currency format
        if "currency_style" not in style_names:
            currency_style = NamedStyle(name="currency_style")
            currency_style.number_format = '"$"#,##0.00_);("$"#,##0.00)'
            currency_style.alignment = Alignment(horizontal='center', vertical='center')
            workbook.add_named_style(currency_style)
        self.currency_style = "currency_style"

        # Currency style for negative numbers
        if "currency_negative_style" not in style_names:
            currency_negative_style = NamedStyle(name="currency_negative_style")
            currency_negative_style.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            currency_negative_style.alignment = Alignment(horizontal='center', vertical='center')
            workbook.add_named_style(currency_negative_style)
        self.currency_negative_style = "currency_negative_style"

        # Text style
        if "text_style" not in style_names:
            text_style = NamedStyle(name="text_style")
            text_style.alignment = Alignment(horizontal='center', vertical='center')
            workbook.add_named_style(text_style)
        self.text_style = "text_style"

    def update_jacks_buckets(self):
        """Updates Jacks Buckets with recent transactions from Debit Transactions excel file"""
        # Run updateMyBuckets to fetch latest transactions
        self._log("\n🔹 Updating transaction list")
        fetcher = updateMyBuckets.updateMyBuckets()
        fetcher.run()

        # Check if Jacks Buckets sheet exists
        if "Jacks Buckets" not in self.wb.sheetnames:
            self._log("⚠ Jacks Buckets sheet not found.", is_error=True)
            return

        bucket_sheet = self.wb["Jacks Buckets"]
        
        # Find the last transaction date in Jacks Buckets
        last_date = None
        for row in range(2, bucket_sheet.max_row + 1):  # Assuming row 1 is header
            date_cell = bucket_sheet[f'A{row}'].value
            if date_cell:
                try:
                    if isinstance(date_cell, datetime):
                        # If it's already a datetime object, use it directly
                        last_date = date_cell
                    elif isinstance(date_cell, str):
                        # If it's a string, parse it
                        last_date = datetime.strptime(date_cell, "%d/%m/%Y")
                except ValueError:
                    continue
        
        if not last_date:
            self._log("⚠ No valid dates found in Jacks Buckets.", is_error=True)
            return

        self._log(f"✅ Last transaction date in Jacks Buckets: {last_date.strftime('%d/%m/%Y')}")

        # Find the latest Debit Transactions file
        debit_file = None
        for file in os.listdir(TRANSACTION_DIRECTORY):
            if file.startswith("Debit Transactions") and file.endswith(".xlsx"):
                debit_file = os.path.join(TRANSACTION_DIRECTORY, file)
                break

        if not debit_file:
            self._log("⚠ No Debit Transactions file found.", is_error=True)
            return

        self._log(f"✅ Found Debit Transactions file: {debit_file}")

        # Load the Debit Transactions workbook
        try:
            debit_wb = load_workbook(debit_file)
            debit_sheet = debit_wb.active
        except Exception as e:
            self._log(f"❌ Error loading Debit Transactions file: {str(e)}", is_error=True)
            return

        # Process new transactions
        new_transactions = []
        for row in range(2, debit_sheet.max_row + 1):  # Assuming row 1 is header
            date_str = debit_sheet[f'A{row}'].value
            if not date_str:
                continue

            try:
                if isinstance(date_str, datetime):
                    trans_date = date_str
                elif isinstance(date_str, str):
                    # Convert YYYY-MM-DD to datetime
                    trans_date = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    continue

                if trans_date > last_date:
                    description = debit_sheet[f'B{row}'].value
                    amount = debit_sheet[f'C{row}'].value

                    # Categorization rules (hardcoded)
                    category = None
                    if description.startswith("Direct Credit 617702") and "PAYPAL AUSTRALIA" in description:
                        category = "DataAnnotation"
                    elif "Salary" in description:
                        category = "Salary"
                    elif "Jack weekly spend" in description:
                        category = "Salary"
                    elif "Solar Loan" in description:
                        category = "Salary"
                    elif "Transfer to xx9545" in description:
                        category = "Salary"

                    new_transactions.append({
                        'date': trans_date,
                        'description': description,
                        'category': category,
                        'amount': amount
                    })
                    self._log(f"Jacks Buckets updated successfully.")
            except ValueError:
                continue

        self.setup_styles(self.wb)

        # Add new transactions to Jacks Buckets
        if new_transactions:
            # Sort transactions by date (oldest first) before adding
            new_transactions.sort(key=lambda x: x['date'])
            
            start_row = bucket_sheet.max_row + 1
            for i, trans in enumerate(new_transactions, start=start_row):
                # Date column (A)
                date_cell = bucket_sheet[f'A{i}']
                date_cell.value = trans['date']  # Set as datetime object
                date_cell.style = self.date_style

                # Description column (B)
                desc_cell = bucket_sheet[f'B{i}']
                desc_cell.value = trans['description']
                desc_cell.style = self.text_style

                # Category column (C)
                cat_cell = bucket_sheet[f'C{i}']
                cat_cell.value = trans['category']
                cat_cell.style = self.text_style

                # Amount column (D)
                amount_cell = bucket_sheet[f'D{i}']
                amount_cell.value = trans['amount']
                # Apply appropriate currency style based on value
                if trans['amount'] < 0:
                    amount_cell.style = self.currency_negative_style
                else:
                    amount_cell.style = self.currency_style

                # Copy formatting from the row above
                if start_row > 2:
                    for col in ['A', 'B', 'C', 'D']:
                        above_cell = bucket_sheet[f'{col}{i-1}']
                        current_cell = bucket_sheet[f'{col}{i}']
                        if above_cell.fill and above_cell.fill != PatternFill():
                            current_cell.fill = copy(above_cell.fill)
                        if above_cell.border:
                            current_cell.border = copy(above_cell.border)

            self._log(f"✅ Added {len(new_transactions)} new transactions to Jacks Buckets")
        else:
            self._log("ℹ No new transactions found since last update", is_error=True)

    def save_workbook(self, output_path):
        """Saves the modified workbook."""
        try:
            self.wb.save(output_path)
            self._log(f"\n✅ Saved to {output_path}")
        except Exception as e:
            self._log(f"❌ Error saving workbook: {str(e)}", is_error=True)

    def run_all_updates(self):
        """Run all update operations in sequence"""
        self.convert_previous_month_to_values()
        self.update_formulas()
        self.update_jacks_buckets()

if __name__ == "__main__":
    try:
        updater = BudgetUpdater(
            file_path='summary_updated.xlsm',
            workbook_name="2025 Monthly Spend.xlsx",
            verbose=False  # Set to True for detailed logs
        )
        updater.run_all_updates()
        updater.save_workbook('summary_updated.xlsm')
    except Exception as e:
        print(f"❌ An error occurred: {str(e)}")