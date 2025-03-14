import requests
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import calendar
import numpy as np
import os
from config import (  # Import settings from config.py
    POCKETSMITH_API_KEY,
    ULTIMATE_AWARDS_CC_ID,
    PEOPLE,
    TRANSACTION_DIRECTORY,
    RUBY_FILL,
    JACK_FILL,
    BOTH_FILL,
)

'''
This program will fetch all credit card debits & credits starting from the date in yyyy-mm-dd format as input and ending with today's date as the end date. It will then parse this data into an excel spreadsheet named 'MMM Week X - 202X.xlsx'. As part of this data, it will leverage built-in categorizations from PocketSmith. 

Additionally, labels are added sparringly to identify whether a transaction is made by Ruby Jack or Both. This is so that in the summary table that is added at the right hand side of this transaction list, will summarize all the spend that is made by Jack & Ruby. 

Currently, manual analysis of the data still needs to be made to identify labels properly. Once this is identified, the spreadsheet can then be shared across to pay off the credit card based on the amount specified in the summary table.

TODO:
-Simplify the assignment of categories for all transactions. Maybe still stick to the ones used in Apple Numbers - DONE ISH - Decided to use categories from PocketSmith itself for more granularity.
-Transfer all existing categorizations to the the new simplified categories
--> Need to ask myself why I want to simplify. Is it really necessary? Isn't more granularity the better?
--> Maybe simplify "Bills" only? But it's good to know breakdown of water, electricity, gas bills on a per annum basis instead of bundling it all under "Bills".

26/12/24: Consider changing the "Both" label name to something else which can suit 3+ PEOPLE added. 28/12/24: Plan was to chang label naming such that if len(PEOPLE) >= 3, then change label naming to "All", else set it to "Both". However, found that the remaining solution on label naming & color filling is also hardcoded to either jack, ruby or both names. Hence this fix would require a change on how all of these variables are leveraged. a.k.a fixing this is more complicated than expected.
26/12/24: Add code to GitHub but use gitignore on config.py file.

-Consider how I can append all weekly transactions into one spreadsheet for a given Month i.e., 1 spreadsheet, 1 sheet, 4-5 weeks of transactions per sheet. - DONE
-Consider how I can append all monthly transactions into one spreadsheet for a given year i.e., 1 spreadsheet, 12 sheets, 4-5 weeks of transactions in each sheet. - DONE


-Consider how I can add Jacks Buckets in a separate spreadsheet somewhere? WIP
--Need to leverage the debits account here.
-Consider how I can add the "Total Balance" offset balance in a separate spreadsheet (the same one as above?). Also need to consider how I can automate this as well. The idea is if my weekly spend changes, savings are added or bills are included, then all of this needs to be dynamically updated as well. Current process is manual. - DONE
--Will need to add offset account here. 

-Consider turning this into a class. - WIP

'''

# Dynamically generate the filename based on the current week, month, and year
def generate_spreadsheet_name(start_date,TRANSACTION_DIRECTORY):
    # Parse the start date
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()

    # Use today's date as the end date
    end_date_obj = datetime.now().date()

    # Format the dates into DD-MM format with the month name
    start_date_str = start_date_obj.strftime('%d')  # E.g., "20"
    end_date_str = end_date_obj.strftime('%d')
    month_name = end_date_obj.strftime('%b')  # E.g., "Dec"
    year = end_date_obj.year
    
    today = datetime.now()
    week = week_of_month(today.strftime('%Y-%m-%d'))  # Week number in the month

    # Format the filename as required
    filename = f"{start_date_str}-{end_date_str} {month_name} Week {week} - {year}.xlsx"
    
    return f"{TRANSACTION_DIRECTORY}{filename}"

# Function to fetch bank feed transactions from PocketSmith API
def fetch_transactions(start_date=None):
    page = 1
    all_transactions = []

    while True:
        url = f"https://api.pocketsmith.com/v2/accounts/{ULTIMATE_AWARDS_CC_ID}/transactions?page={page}&start_date={start_date}&end_date={datetime.now()}"
        
        headers = {"accept": "application/json", "X-Developer-Key": POCKETSMITH_API_KEY}

        print(f"Fetching transactions from PocketSmith API - Page {page}...")
        response = requests.get(url, headers=headers)

        if response.status_code != 200:
            #print(f"Error fetching transactions: {response.status_code} {response.text}")
            return all_transactions

        transactions = response.json()

        if not transactions:  # If there are no transactions, we stop fetching
            break
        
        # Parse transactions to match our format
        for tx in transactions:
            if tx.get('category') and 'title' in tx['category']:
                category_title = tx['category']['title']
            else:
                category_title = ""
            all_transactions.append({
                'date': tx['date'],
                'description': tx['payee'],
                'bank_category': category_title,
                'amount': tx['amount']
            })

        page += 1  # Move to the next page for the next iteration
    return all_transactions


# Function to auto-label bank categories
# TODO: Fix this so it is not hardcoded and is dependent on the PEOPLE labels. For now, it is currently hardcoded.
def auto_label_bank_category(bank_category):
    # Return specific labels based on bank category
    if not bank_category or bank_category in ["Dining", "Travel"]:
        return None  # No label for these categories
    elif bank_category in ["Recreation", "Professional Services"]:
        return "Jack"  # Return "Jack" for these categories
    else:
        return "Both"  # Default label for all other categories


# Function to categorize and label transactions
def categorize_and_label_transactions(transactions):
    categorized_transactions = []
    
    for tx in transactions:
        description = tx['description']
        amount = tx['amount']
        bank_category = tx['bank_category']

        # Apply auto-labeling to the bank category
        label = auto_label_bank_category(bank_category)

        # Example categorization based on description (could use PocketSmith rules here)
        category = None
        label = label if label else None  # If label is None, keep as is

        # Add transaction with categorization and label
        categorized_transactions.append({
            'Date': tx['date'],
            'Description': description,
            'Amount': amount,
            'Category': category,
            'Bank Category': tx['bank_category'],  # Add bank_category to the data
            'Label': label
        })
    
    return categorized_transactions

def add_summary_table(ws, data, last_row):
    # Find the index of "Amount" and "Label" columns dynamically by searching headers
    headers = [cell.value for cell in ws[1]]  # Get all headers from the first row
    
    # Search for columns based on header names (case-insensitive search)
    amount_col_idx = headers.index("Amount") + 1  # Adding 1 because openpyxl is 1-indexed
    label_col_idx = headers.index("Label") + 1  # Adding 1 because openpyxl is 1-indexed

    # Define the range for valid rows (from row 2 to the last row of data)
    first_data_row = 2
    last_data_row = last_row
    amount_range = f"{openpyxl.utils.get_column_letter(amount_col_idx)}{first_data_row}:{openpyxl.utils.get_column_letter(amount_col_idx)}{last_data_row}"
    label_range = f"{openpyxl.utils.get_column_letter(label_col_idx)}{first_data_row}:{openpyxl.utils.get_column_letter(label_col_idx)}{last_data_row}"

    # Find the next available column (after the last column of the data) and shift it to the right by 1
    last_col = len(headers) + 2  # Shift by 1 to place summary table in a new column
    
    # Dynamically calculate the number of users
    num_users = len(PEOPLE)

    # Dynamically add rows for each person in PEOPLE
    row_idx = 3  # Start from row 3 for the first person
    for person in PEOPLE:
        # Add the person's name
        ws.cell(row=row_idx, column=last_col, value=person)

        # Add formula for the person's total amount
        ws.cell(
            row=row_idx,
            column=last_col + 1,
            value=f'=SUMIFS({amount_range}, {label_range}, "{person}") + SUMIFS({amount_range}, {label_range}, "Both") / {num_users}'
        )

        # Apply color formatting based on the person's index (optional customization)
        if person == "Jack":  # Example custom color for Jack
            fill_color = JACK_FILL
        elif person == "Ruby":  # Example custom color for Ruby
            fill_color = RUBY_FILL
        else:  # Default color for additional names
            fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # TODO: Expand unique colors for X number of PEOPLE added. Currently, this is hardcoded such that any users past 3 will be set to Yellow only.

        ws.cell(row=row_idx, column=last_col).fill = fill_color  # Name cell color
        ws.cell(row=row_idx, column=last_col + 1).fill = fill_color  # Amount cell color

        # Increment the row index for the next person
        row_idx += 1

    # Formula for "Total Weekly Spend" (added at the next row after PEOPLE)
    ws.cell(row=row_idx, column=last_col, value="Total Weekly Spend")
    ws.cell(
        row=row_idx,
        column=last_col + 1,
        value=f'=SUM({openpyxl.utils.get_column_letter(last_col + 1)}3:{openpyxl.utils.get_column_letter(last_col + 1)}{row_idx - 1})'
    )

    # Apply formatting to all cells in the summary table
    for row in range(3, row_idx + 1):  # Rows for all people + Total Weekly Spend
        for col in range(last_col, last_col + 2):  # Columns for names and amounts
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal='center')
            
            # Apply thin border to all cells
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border
            
            # Format Amount column as currency
            if col == last_col + 1:
                cell.number_format = '"$"#,##0.00'

    # Auto-fit column widths for the summary table
    #max_length_label = max(len(person) for person in PEOPLE) + 2  # Find the longest name
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col)].width = 18.57 # Hardcoded currently. Can be amended to dynamically change based on current column length +2, but this may have a performance impact.
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col + 1)].width = 13.57  # Approx. width for amount column

    print("Summary table added successfully.")

# Function to save data to Excel with formatting, dropdown lists, AutoFit, and borders
def save_to_excel(data):

    # Remove credits that don't have matching debits
    # filtered_data, credits = remove_non_matching_credits(data)

    # 30th Jan 2025: All credits now included by default
    filtered_data = data

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Add "Week X" above the data table based on the end_date
    week_number = week_of_month(datetime.now().strftime('%Y-%m-%d'))
    week_label = f"Week {week_number}"


    # Set column headers (swapped Amount and Category columns)
    headers = [f'Date - {week_label}', 'Description', 'Amount', 'Category', 'Bank Category', 'Label']  # Added 'Bank Category' column
    ws.append(headers)
    
    # Apply header formatting (bold and color)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True, color="000000")  # Black font color
    
    # Apply header style
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data rows with Amount and Category swapped
    for row in filtered_data:
        # Swapping Category and Amount columns: [Date, Description, Amount, Category, Label]
        ws.append([row['Date'], row['Description'], row['Amount'], row['Category'], row['Bank Category'], row['Label']])
 
    # Create date style
    date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")
    wb.add_named_style(date_style)
    
    # Format the Date column (Column 1) as a date
    for row in range(2, len(filtered_data) + 2):
        ws[f"A{row}"].style = date_style

    # Format the Amount column (Column 3) as currency
    for row in range(2, len(filtered_data) + 2):
        ws[f"C{row}"].number_format = '"$"#,##0.00'

    # Add dropdown for Category column (Column 4)
    category_list = ["Home", "Entertainment", "Dining", "Personal Items", "Medical", "Vehicle", "Travel", "Other", 
        "Savings", "Mortgage", "Bills", "Gifts", "Groceries", "Subscription"]
    category_validation = DataValidation(type="list", formula1=f'"{",".join(category_list)}"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(category_validation)

    # Apply the category validation to the Category column (Column D) for all rows
    for row in range(2, len(filtered_data) + 2):
        category_validation.add(ws[f"D{row}"])

    # Add dropdown for Label column (Column 6)
    label_list = PEOPLE + ["Both"]  # Combine PEOPLE with the "Both" option dynamically
    label_validation = DataValidation(type="list", formula1=f'"{",".join(label_list)}"', allow_blank=True, showDropDown=False)
    label_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(label_list)}"',  # Dynamically create the dropdown from PEOPLE and "Both"
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(label_validation)
    
    # Apply the label validation to the Label column (Column E) for all rows
    last_row = len(filtered_data) + 1  # Last row with actual data
    for row in range(2, last_row + 1):
        label_validation.add(ws[f"F{row}"])

    # AutoFit column widths
    for col in range(1,len(filtered_data[0])):  # Columns A to F (1 to 6)
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)  # Get column name (A, B, C, etc.)
        
        for row in range(1, len(filtered_data) + 2):  # Including header row
            try:
                cell_value = str(ws[f"{column}{row}"].value)
                max_length = max(max_length, len(cell_value))
            except:
                pass
        adjusted_width = (max_length + 3)  # Adding a little extra space
        ws.column_dimensions[column].width = adjusted_width

    # Apply borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply borders to all rows and columns with data
    for row in ws.iter_rows(min_row=1, max_row=len(filtered_data) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Add Conditional Formatting for Row Highlighting
    # TODO: Remove hardcoding here for formulas
    last_row = len(filtered_data) + 1 # TODO: remove. Already defined previously
    ws.conditional_formatting.add(
        f"A2:F{last_row}",
        FormulaRule(formula=['$F2="Ruby"'], fill=RUBY_FILL)
    )
    ws.conditional_formatting.add(
        f"A2:F{last_row}",
        FormulaRule(formula=['$F2="Jack"'], fill=JACK_FILL)
    )
    ws.conditional_formatting.add(
        f"A2:F{last_row}",
        FormulaRule(formula=['$F2="Both"'], fill=BOTH_FILL)
    )

    # Add the summary table at the right
    add_summary_table(ws, filtered_data, last_row)

    # Apply center alignment to all titles (row 1) and all remaining rows with values
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            # Apply center alignment to row 1 (titles) and all rows except column B from row 2 onwards
            if cell.row == 1 or (cell.column != 2 and cell.row > 1):  
                cell.alignment = Alignment(horizontal='center')


    # Save the workbook to file
    wb.save(SPREADSHEET_PATH)
    print(f"Data saved to {SPREADSHEET_PATH}")
    
#@staticmethod
def week_of_month(date_str):
    # Set the first weekday to be Monday
    calendar.setfirstweekday(calendar.MONDAY)
    
    # Parse the input date string
    date = datetime.strptime(date_str, '%Y-%m-%d')
    year, month, day = date.year, date.month, date.day
    
    # Create a month calendar array
    month_calendar = np.array(calendar.monthcalendar(year, month))
    
    # Find the week of the month
    week_of_month = np.where(month_calendar == day)[0][0] + 1
    return week_of_month

def get_last_run_date():
    """Get the date when this program was last run."""
    last_run_file = os.path.join(TRANSACTION_DIRECTORY, 'last_run.txt')
    try:
        with open(last_run_file, 'r') as f:
            last_run_date = f.read().strip()
            return datetime.strptime(last_run_date, '%Y-%m-%d').date()
    except (FileNotFoundError, ValueError):
        return None

def save_last_run_date():
    """Save the current date as the last run date."""
    last_run_file = os.path.join(TRANSACTION_DIRECTORY, 'last_run.txt')
    with open(last_run_file, 'w') as f:
        f.write(datetime.now().strftime('%Y-%m-%d'))

def get_user_start_date():
    """Get start date from user input with validation."""
    last_run = get_last_run_date()
    if last_run:
        print(f"\nLast time this program was run: {last_run.strftime('%Y-%m-%d')}")
    
    while True:
        start_date = input("\nPlease enter the start date (YYYY-MM-DD): ").strip()
        try:
            # Validate date format
            datetime.strptime(start_date, '%Y-%m-%d')
            return start_date
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD format.")

# Main function to fetch, categorize, and save transactions
def main(start_date=None):
    """Main function to fetch, categorize, and save transactions."""
    if start_date is None:
        start_date = get_user_start_date()
    
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.now().date()
    
    days_span = (end_date_obj - start_date_obj).days + 1
    print(f"\nFetching transactions starting from {start_date_obj.strftime('%dth %B %Y')} to {end_date_obj.strftime('%dth %B %Y')} (spans {days_span} days).")

    transactions = fetch_transactions(start_date)
    categorized_transactions = categorize_and_label_transactions(transactions)
    
    # Generate spreadsheet filename with updated logic
    global SPREADSHEET_PATH
    SPREADSHEET_PATH = generate_spreadsheet_name(start_date, TRANSACTION_DIRECTORY)
    
    save_to_excel(categorized_transactions)
    save_last_run_date()  # Save the current date as last run date


# Run the main function with a specified start_date for testing
if __name__ == "__main__":
    main()
