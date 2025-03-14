import os
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from copy import copy
from openpyxl.formula.translate import Translator
from config import (
    SPREADSHEET_DIRECTORY,
    TRANSACTION_DIRECTORY,
    CURRENT_YEAR,
    MASTER_SPREADSHEET_NAME,
    RUBY_FILL,
    JACK_FILL,
    BOTH_FILL,
    BACKUP_DIRECTORY,
)


class SpreadsheetCollator:
    """Class to manage the collation of weekly spreadsheets into monthly sheets."""
    
    # Class constants
    MAX_COLUMN = 9  # Up to column I
    MONTH_ABBREVIATIONS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def __init__(self, verbose=True):
        """Initialize the collator with configuration settings and verbosity option."""
        self.master_wb = None
        self.data_appended = False
        self.verbose = verbose

    def _log(self, message):
        """Helper method to handle conditional printing based on verbosity setting."""
        if self.verbose:
            print(message)

    def _copy_cell_with_styles(self, source_cell, target_cell):
        """Helper method to copy cell value and styles, handling formulas appropriately."""
        if source_cell.data_type == 'f':
            # Translate formula to new cell position
            target_cell.value = Translator(source_cell.value, 
                                        origin=source_cell.coordinate
                                        ).translate_formula(target_cell.coordinate)
        else:
            target_cell.value = source_cell.value

        # Copy cell styles if present
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def _copy_column_widths(self, source_ws, target_ws):
        """Helper method to copy column widths from source to target worksheet."""
        for col_idx in range(1, self.MAX_COLUMN + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_ws.column_dimensions:
                source_width = source_ws.column_dimensions[col_letter].width
                if source_width:
                    target_ws.column_dimensions[col_letter].width = source_width
        
        # Set column E width to 24 for all collated sheets
        target_ws.column_dimensions['E'].width = 24
        self._log("Set column E width to 24")

    def _copy_conditional_formatting(self, source_ws, target_ws):
        """Helper method to copy and update conditional formatting rules."""
        for rule in source_ws.conditional_formatting:
            original_cf = source_ws.conditional_formatting[rule]
            for applies_to_formula in rule:
                if applies_to_formula[0] == 'sqref':
                    new_applies_to_formula = 'A2:F300'
                    del source_ws.conditional_formatting
                    
                    # Create a mapping for faster lookup
                    fill_map = {
                        '$F2="Ruby"': RUBY_FILL,
                        '$F2="Jack"': JACK_FILL,
                        '$F2="Both"': BOTH_FILL
                    }
                    
                    for cf_rule in original_cf:
                        original_formula = cf_rule.formula[0] if cf_rule.formula else None
                        fill = fill_map.get(original_formula, BOTH_FILL)
                        target_ws.conditional_formatting.add(
                            new_applies_to_formula,
                            FormulaRule(formula=[original_formula], fill=fill)
                        )
                else:
                    print("Error: sqref not found")
        self._log("Copied and updated conditional formatting successfully.")

    def copy_data_with_format_and_conditional_formatting(self, source_ws, target_ws, 
                                                       start_row, copy_cf=True):
        """
        Copy data, formatting, and conditional formatting from source to target worksheet.
        
        Args:
            source_ws: Source worksheet
            target_ws: Target worksheet
            start_row: Starting row in target worksheet
            copy_cf: Whether to copy conditional formatting (default: True)
        """
        # Copy data and styles
        for row in source_ws.iter_rows(min_row=1, max_col=self.MAX_COLUMN):
            for cell in row:
                target_cell = target_ws.cell(
                    row=start_row + cell.row - 1,
                    column=cell.column
                )
                self._copy_cell_with_styles(cell, target_cell)

        # Copy column widths
        self._copy_column_widths(source_ws, target_ws)

        # Copy conditional formatting if specified
        if copy_cf:
            self._copy_conditional_formatting(source_ws, target_ws)

    def ensure_master_workbook_exists(self):
        """Ensure the master workbook exists, creating it if necessary."""
        master_path = os.path.join(SPREADSHEET_DIRECTORY, MASTER_SPREADSHEET_NAME)
        if not os.path.exists(master_path):
            self.master_wb = Workbook()
            default_sheet = self.master_wb.active
            default_sheet.title = "Default"
            self.master_wb.save(master_path)
            self._log(f"Created new master spreadsheet: {MASTER_SPREADSHEET_NAME}")
        else:
            self._log(f"Master spreadsheet found: {MASTER_SPREADSHEET_NAME}")

    def backup_existing_spreadsheet(self):
        """Backup the existing master spreadsheet if it exists by moving it to backup directory."""
        master_path = os.path.join(SPREADSHEET_DIRECTORY, MASTER_SPREADSHEET_NAME)
        
        if os.path.exists(master_path):
            # Generate backup filename with current date
            current_date = datetime.now().strftime("%Y-%m-%d")
            filename, extension = os.path.splitext(MASTER_SPREADSHEET_NAME)
            backup_filename = f"{filename}_{current_date}{extension}"
            backup_path = os.path.join(BACKUP_DIRECTORY, backup_filename)
            
            # Ensure backup directory exists
            os.makedirs(BACKUP_DIRECTORY, exist_ok=True)
            
            # Move the file to backup directory instead of copying
            try:
                shutil.move(master_path, backup_path)
                self._log(f"Moved existing spreadsheet to backup: {backup_path}")
                return True
            except Exception as e:
                print(f"Error backing up spreadsheet: {e}")
                return False
        return False

    def _process_file(self, file, month_name, year):
        """Process a single weekly file and append it to the appropriate monthly sheet."""
        full_month_name = datetime.strptime(month_name, "%b").strftime("%B")
        self._log(f"Processing file: {file} -> Month: {full_month_name}")

        try:
            weekly_wb = load_workbook(os.path.join(TRANSACTION_DIRECTORY, file))
            weekly_ws = weekly_wb.active
        except Exception as e:
            print(f"Error reading file {file}: {e}")  # Always print exceptions
            return

        if full_month_name not in self.master_wb.sheetnames:
            self.master_wb.create_sheet(title=full_month_name)
        month_ws = self.master_wb[full_month_name]

        # Determine start row more efficiently
        start_row = 1
        if month_ws.max_row > 1:
            start_row = month_ws.max_row + 1
            if month_ws.cell(row=start_row - 1, column=1).value:
                start_row += 1

        self._log(f"Appending data to {full_month_name} starting at row {start_row}.")
        self.copy_data_with_format_and_conditional_formatting(
            weekly_ws, month_ws, start_row, copy_cf=(start_row == 1)
        )
        self.data_appended = True

    def collate_monthly_spreadsheets(self):
        """Collate all weekly spreadsheets into monthly sheets in the master workbook."""
        self.ensure_master_workbook_exists()
        
        # Backup existing spreadsheet before making changes
        self.backup_existing_spreadsheet()
        
        try:
            # Create a new workbook since the old one was moved to backup
            self.master_wb = Workbook()
            default_sheet = self.master_wb.active
            default_sheet.title = "Default"
            self._log(f"Created new master workbook after backing up the previous one")
        except Exception as e:
            print(f"Error creating new master workbook: {e}")
            return

        # Get and filter files in one pass
        weekly_files = []
        try:
            for f in os.listdir(TRANSACTION_DIRECTORY):
                if f.endswith(".xlsx") and "Week" in f:
                    try:
                        week_num = int(f.split("Week")[1].split()[0])
                        weekly_files.append((f, week_num))
                    except (ValueError, IndexError):
                        self._log(f"Skipping file {f}: cannot parse week number.")
                        continue
        except Exception as e:
            print(f"Error accessing transaction directory: {e}")
            return
            
        # Sort files by week number
        weekly_files.sort(key=lambda x: x[1])
        weekly_files = [f[0] for f in weekly_files]

        for file in weekly_files:
            try:
                month_name = next((m for m in self.MONTH_ABBREVIATIONS if m in file), None)
                if not month_name:
                    self._log(f"Skipping file {file}: month abbreviation not found.")
                    continue

                year_part = file.split(" ")[-1].replace(".xlsx", "")
                if not year_part.isdigit() or int(year_part) != CURRENT_YEAR:
                    self._log(f"Skipping file {file}: invalid or mismatched year.")
                    continue

                self._process_file(file, month_name, year_part)
            except Exception as e:
                print(f"Skipping file {file}: error processing ({e})")
                continue

        # Remove Default sheet only if data was appended and there are other sheets
        if self.data_appended and 'Default' in self.master_wb.sheetnames and len(self.master_wb.sheetnames) > 1:
            del self.master_wb['Default']
            self._log("Removed unused Default sheet as data was appended.")

        try:
            # Save the output file to SPREADSHEET_DIRECTORY
            output_path = os.path.join(SPREADSHEET_DIRECTORY, MASTER_SPREADSHEET_NAME)
            self.master_wb.save(output_path)
            self._log(f"All data collated into {output_path}.")
        except Exception as e:
            print(f"Error saving master spreadsheet: {e}")  # Always print exceptions

if __name__ == "__main__":
    collator = SpreadsheetCollator(verbose=False)
    collator.collate_monthly_spreadsheets()