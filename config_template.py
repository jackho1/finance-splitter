# Configuration for the program
from datetime import datetime
from openpyxl.styles import PatternFill
import os

# PocketSmith API settings
# UPDATE THIS
POCKETSMITH_API_KEY = 'YOUR_POCKETSMITH_API_KEY'
#POCKETSMITH_USER_ID = 'YOUR_POCKETSMITH_USER_ID' #not in use currently
ULTIMATE_AWARDS_CC_ID = 'YOUR_ULTIMATE_AWARDS_CC_ID' # used in bank_feeds.py to update weekly transactions
DEBIT_ID = 'YOUR_DEBIT_ID' # used in BudgetUpdater.py to update debit transactions


# List of users to split finance payments with 
PEOPLE = ["Jack", "Ruby"]

# List of colorings used for above users
#TODO: This is stil hardcoded. Need to update all references in all python files if this name is updated below.
RUBY_FILL = PatternFill(start_color="FF2C55", end_color="FF2C55", fill_type="solid")  # Red 
JACK_FILL = PatternFill(start_color="5582AE", end_color="5582AE", fill_type="solid")  # Blue
BOTH_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

# File path for saving spreadsheets
# UPDATE THIS
SPREADSHEET_DIRECTORY = "your/path/here"

# File names
SUMMARY_FILE = "summary_updated.xlsm"  # Master file with VBA and summary sheets

# File path for collating spreadsheets together on a monthly basis. TODO: Fix naming
# Configuration
CURRENT_YEAR = datetime.now().year  # Automatically fetch the current year

TRANSACTION_DIRECTORY = os.path.join(
    SPREADSHEET_DIRECTORY,
    f"{CURRENT_YEAR} Transactions/"
)

# Ensure the transaction directory exists
if not os.path.exists(TRANSACTION_DIRECTORY):
    os.makedirs(TRANSACTION_DIRECTORY)

# Backup directory for storing backup files
BACKUP_DIRECTORY = os.path.join(
    SPREADSHEET_DIRECTORY,
    "Backup/"
)
# Ensure the backup directory exists
if not os.path.exists(BACKUP_DIRECTORY):
    print("Back up directory created")
    os.makedirs(BACKUP_DIRECTORY)

MASTER_SPREADSHEET_NAME = f"{CURRENT_YEAR} Monthly Spend.xlsx"  # Dynamic name based on the year

# Other IDs (if needed in the future)
#INSTITUTION_ID = 'YOUR_INSTITUTION_ID'

# Database configuration
DB_CONFIG = {
    'dbname': 'transactions',
    'user': 'postgres',
    'password': 'password',
    'host': 'localhost',
    'port': '5432'
}

# Transaction fetching configuration
DAYS_TO_FETCH = 30  # Number of days to fetch transactions for

