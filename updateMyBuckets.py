import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Import configuration variables
from config import (
    DEBIT_ID, 
    POCKETSMITH_API_KEY, 
    TRANSACTION_DIRECTORY,
    JACK_FILL
)

class updateMyBuckets:
    def __init__(self):
        """Initialize the class with API key and base URL for PocketSmith."""
        self.api_key = POCKETSMITH_API_KEY
        self.base_url = "https://api.pocketsmith.com/v2"
        self.headers = {
            "Authorization": f"Key {self.api_key}",
            "Accept": "application/json"
        }
        # Define border style for cells
        self.cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def fetch_transactions(self):
        """
        Fetch all debit transactions from PocketSmith API.
        Returns a list of transaction data.
        """
        try:
            url = f"{self.base_url}/accounts/{DEBIT_ID}/transactions"
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            transactions = response.json()
            return transactions
        except requests.RequestException as e:
            print(f"Error fetching transactions: {e}")
            return []

    def format_transaction_data(self, transactions):
        """
        Format the raw transaction data into a structured list for Excel export.
        Returns a list of dictionaries with required fields.
        """
        formatted_data = []
        for transaction in transactions:
            formatted_data.append({
                "Date": transaction.get("date", ""),
                "Description": transaction.get("payee", ""),
                "Amount": transaction.get("amount", 0),
                "Category": transaction.get("category", {}).get("title", "Uncategorized")
            })
        return formatted_data

    def create_excel_file(self, transaction_data):
        """
        Create and format an Excel file with the transaction data.
        Saves the file in the specified TRANSACTION_DIRECTORY.
        """
        # Create output directory if it doesn't exist
        if not os.path.exists(TRANSACTION_DIRECTORY):
            print("Creating a new directory for the output file because it does not exist...")
            os.makedirs(TRANSACTION_DIRECTORY)

        # Generate filename with current date in "DD Mon YY" format
        current_date = datetime.now().strftime("%Y")
        output_file = os.path.join(TRANSACTION_DIRECTORY, f"Debit Transactions {current_date}.xlsx")

        # Convert data to DataFrame
        df = pd.DataFrame(transaction_data)

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Debit Transactions"

        # Define header style
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)

        # Write headers
        headers = ["Date", "Description", "Amount", "Category"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.cell_border

        # Write data rows
        for row_num, row_data in enumerate(transaction_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                if header == "Amount":
                    # Set the actual numeric value
                    amount = float(row_data[header])
                    cell.value = amount
                    # Use Excel's standard Currency format with 2 decimal places
                    cell.number_format = '$#,##0.00;- $#,##0.00'
                    # If amount is negative, explicitly set font color to C00000
                    if amount < 0:
                        cell.font = Font(color="C00000")
                else:
                    cell.value = row_data[header]

                # Apply JACK_FILL from config to all data cells
                cell.fill = JACK_FILL
                # Add border to all data cells
                cell.border = self.cell_border
                
                # Set alignment: center for all columns except Description (column B)
                if col_num == 2:  # Column B (Description)
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    # For currency cells, add extra width to accommodate formatting
                    cell_value = str(cell.value)
                    if column_letter == 'C':  # Amount column
                        # Approximate width considering negative values
                        cell_value = f"${cell_value}" if float(cell.value) >= 0 else f"- ${abs(float(cell_value))}"
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        try:
            wb.save(output_file)
            print(f"Transactions exported successfully to {output_file}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def run(self):
        """
        Main method to orchestrate the transaction fetching and Excel export process.
        """
        raw_transactions = self.fetch_transactions()
        if raw_transactions:
            formatted_data = self.format_transaction_data(raw_transactions)
            if formatted_data:
                self.create_excel_file(formatted_data)
            else:
                print("No transactions found to export.")
        else:
            print("Failed to fetch transactions.")

"""
if __name__ == "__main__":
    fetcher = updateMyBuckets()
    fetcher.run()
"""